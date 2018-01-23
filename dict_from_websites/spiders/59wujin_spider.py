#coding:utf-8
import scrapy
import time
from openpyxl import Workbook, load_workbook

#RENDER_HTML_URL = "http://jxdmyx.com:8050/render.html"

class ChinawujinSpider(scrapy.Spider):
    name = "59wujin_spider"
    allowed_domains = ['59wujin.com', 'jxdmyx.com',]
    path_to_write = "/mnt/hgfs/windows_desktop/classification_and_coding/" +\
            "data/dictionary_building/from_shopping_websites/" +\
            "dict_from_59wujin_" + time.strftime("%Y%m%d", time.localtime()) + ".xlsx"

    def __init__(self):
        try:
            wb = load_workbook(self.path_to_write)
        except Exception, e:
            wb = Workbook()
            wb.save(self.path_to_write)

    def start_requests(self):
        list_urls = ['http://www.59wujin.com/sellfl.html',]
	for url in list_urls:
	    yield scrapy.Request(url=url, callback=self.parse)
    
    def parse(self, response):
	for sel in response.xpath("//div[@class='wrap']/div[@class='pBox']")[1:2]:
            c1_name = sel.xpath("h3/text()").extract()[0]
	    for sel2 in sel.xpath("ul/li"):
		c2_name = sel2.xpath("a/text()").extract()[0]
		c2_url = 'http://www.59wujin.com/' + sel2.xpath("a/@href").extract()[0]
#                print c1_name, c2_name, c2_url
            yield scrapy.Request(url=cate_url, callback=self.getListOfKeywords)


    def getListOfKeywords(self, response):
	for 
