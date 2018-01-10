#coding:utf-8
import scrapy
import numpy as np
import time
from openpyxl import Workbook, load_workbook


RENDER_HTML_URL = "http://localhost:8050/render.html"

class TaobaoSpider(scrapy.Spider):
    name = 'taobao_spider'
    start_urls = ['https://www.taobao.com']
    allowed_domains = ['taobao.com', 'localhost',]

    path_to_write = 'data/dict_from_taobao_' +\
            time.strftime("%Y%m%d", time.localtime()) + '.xlsx'

    def __init__(self):
        try:
            wb = load_workbook(self.path_to_write)
        except Exception, e:
            wb = Workbook()
	    wb.save(self.path_to_write)

    def parse(self, response):
        for sel in response.xpath("//ul[@class='service-bd']"):
            list_cate_name = sel.xpath("li/a/text()").extract()
            list_cate_url = sel.xpath("li/a/@href").extract()
	# 有部分链接缺少'https:', 如: ("//mei.taobao.com"
	list_cate_url = ['https:'+u if 'http' not in u else u for u in list_cate_url]
        # (汽车)用品 VS (母婴)用品, 两个用品均可去掉
        dict_cate_url = dict(zip(list_cate_name, list_cate_url))
        dict_cate_url.pop(u'孕产') # 同 童装玩具 而去掉
        dict_cate_url.pop(u'用品') # 同 童装玩具 而去掉
        dict_cate_url.pop(u'数码') # 同 家电 而去掉
        dict_cate_url.pop(u'手机') # 同 家电 而去掉
        dict_cate_url.pop(u'保健品') # 不常见且没有有效关键字 而去掉
        dict_cate_url.pop(u'户外') # 同 运动 而去掉
#        print len(list_cate_name)
#        print len(dict_cate_url.keys())

#        for k in dict_cate_url.keys():
#            if 'http' not in dict_cate_url[k]:
#                print k, dict_cate_url[k]

        list_cate = [
                u'女装', u'男装', u'内衣', u'鞋靴', u'箱包', u'配件',
                u'童装玩具', u'家电', u'美妆', u'洗护',
                u'珠宝', u'眼镜', u'手表', u'运动', u'乐器',]
#                u'游戏', u'动漫', u'影视',
#                u'美食', u'生鲜', u'零食', u'鲜花', u'宠物', u'农资',
#                u'房产', u'装修', u'家具', u'家饰', u'家纺', u'汽车',
#                u'二手车', u'办公', u'DIY', u'五金电子', u'百货', u'货厨',
#                u'家庭保健', u'学习', u'卡券', u'本地服务',]
        list_method = [self.crawlingNvzhuang, self.crawlingNanzhuang,
                self.crawlingNeiyi, self.crawlingXie, self.crawlingXiangbao,
                self.crawlingPei, self.crawlingQbb, self.crawlingTbdc,
                self.crawlingMei, self.crawlingXihuyongpin, self.crawlingZhubao,
                self.crawlingYanjing, self.crawlingShoubiao, self.crawlingCoolcityhome,]
        dict_cate_method = dict(zip(list_cate, list_method))

        # 测试用只取最后一个
        for cate_name in list_cate[-1:]:
            cate_url = dict_cate_url[cate_name]
#            print cate_name, cate_url
            cate_url = RENDER_HTML_URL + "?url=" + cate_url + "&timeout=10&wait=2"
            yield scrapy.Request(url=cate_url, callback=dict_cate_method[cate_name])




    def crawlingCoolcityhome(self, response):
        """运动、户外 淘宝字页面称之为 酷玩城"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'运动'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'运动')

        for sel in response.xpath("//ul[@class='list-name']"):
            list_c1 = sel.xpath("li/a/text()").extract()
            list_c1 = [c.strip() for c in list_c1 if c.strip() != u'']
#            print repr(list_c1).decode('unicode-escape')

        for idx,sel in enumerate(response.xpath("//div[@class='list-k']/ul")):
            list_kws_c1 = sel.xpath("li/a/text()").extract()
            list_kws_c1 = [c.strip() for c in list_kws_c1]
#            print repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, list_c1[idx], kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingShoubiao(self, response):
        """手表"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'手表'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'手表')

        for idx,sel in enumerate(response.xpath("//div[@class='bd']/ul/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = eval(sel.xpath("dl/textarea/text()").extract()[0])['custom']
            list_kws_c1 = [d['cat_name'].decode('utf-8') for d in list_kws_c1]
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingYanjing(self, response):
        """眼镜"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'眼镜'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'眼镜')

        for idx,sel in enumerate(response.xpath("//div[@class='bd']/ul/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = eval(sel.xpath("dl/textarea/text()").extract()[0])['custom']
	    list_kws_c1 = [d['cat_name'].decode('utf-8') for d in list_kws_c1]
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingZhubao(self, response):
        """珠宝"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'珠宝'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'珠宝')

        for idx,sel in enumerate(response.xpath("//div[@class='bd']/ul/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = eval(sel.xpath("dl/textarea/text()").extract()[0])['custom']
            list_kws_c1 = [d['cat_name'].decode('utf-8') for d in list_kws_c1]
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [idx+1, c1, kw,]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)

	
    def crawlingXihuyongpin(self, response):
        """洗护"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'洗护'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'洗护')

        list_c1 = []
        for sel in response.xpath("//div[@class='head-line']"):
            c1 = sel.xpath("h4/text()").extract()[0]
            list_c1.append(c1)
#        print repr(list_c1).decode('unicode-escape')

        for idx,sel in enumerate(response.xpath("//div[@class='list-wrap']")):
            list_kws_c1 = sel.xpath("p/a/text()").extract()
            for kw in list_kws_c1:
                list_to_write = [idx+1, list_c1[idx], kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingMei(self, response):
        """美妆"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'美妆'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'美妆')

        for idx,sel in enumerate(response.xpath(
                "//div[@class='market-wrap clearfix sm-cat-list-main']/dl")):
            # 首页显示类别
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [d['cat_name'].decode('unicode-escape') for d in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
                list_to_write = [idx+1, list_old_cate[0], '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            # 扩展(及隐藏)类别
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_cate = [d['cat_name'].decode('unicode-escape') for d in list_extra_cate]
            list_istitle = [d['is_title']=='true' for d in list_extra_cate]
#            print repr(list_cate).decode('unicode-escape')
#            print list_istitle
            list_index = np.argwhere(np.array(list_istitle)).T[0].tolist()
            list_index.append(len(list_cate))
            list_zip_cate_kw = [(list_index[i], list_index[i+1]) \
                    for i in range(len(list_index)-1)]
            for i,j in list_zip_cate_kw:
                for k in range(i+1,j):
                    list_to_write = [idx+1, list_old_cate[0], list_cate[i], list_cate[k]]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingTbdc(self, response):
        """家电、数码、手机 淘宝字页面称之为 淘宝电场"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'家电'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'家电')

        list_c1 = []
        for sel in response.xpath("//div[@class='nav-p']/p"):
            c1 = sel.xpath("span/text()").extract()[0]
            list_c1.append(c1)
#        print repr(list_c1).decode('unicode-escape')

        # 抓取二级类别及其关键字
        for i,sel in enumerate(response.xpath("//div[@class='nav-text']/div")):
            for sel2 in sel.xpath("div"):
                c2 = sel2.xpath("div/text()").extract()[0]
                list_kws_c2 = sel2.xpath("div/a/text()").extract()
                for kw in list_kws_c2:
                    list_to_write = [i+1, list_c1[i], c2, kw]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingQbb(self, response):
        """童装玩具(淘宝详细页标题 亲宝贝)"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'童装玩具'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'童装玩具')

        # 将返回存入本地文件, 查看网页是否是JS渲染之后的
#        with open('data/test.html', 'w') as f:
#            f.write(response.body)

        for i,sel in enumerate(response.xpath("//ul[@class='nav-lists']/li")):
            list_c1_and_kws = sel.xpath("a/text()").extract() 
#            print repr(list_c1_and_kws).decode('unicode-escape')
            for kw in list_c1_and_kws[1:]:
                list_to_write = [i+1, list_c1_and_kws[0], kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingPei(self, response):
        """配件, 与箱包的解析方式一致, 即下面只是更改了写入excel的表名"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'配件'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'配件')

        for idx,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            # 首页显示类别
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [c['cat_name'].decode('utf-8') for c in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
                list_to_write = [idx+1, list_old_cate[0], '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            # 扩展(及隐藏)类别
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_cate = [d['cat_name'].decode('utf-8') for d in list_extra_cate]
            list_istitle = [d['is_title']=='true' for d in list_extra_cate]
#            print repr(list_cate).decode('unicode-escape')
#            print list_istitle
            list_index = np.argwhere(np.array(list_istitle)).T[0].tolist()
            list_index.append(len(list_cate))
            list_zip_cate_kw = [(list_index[i], list_index[i+1]) \
                    for i in range(len(list_index)-1)]
            for i,j in list_zip_cate_kw:
                for k in range(i+1,j):
                    list_to_write = [idx+1, list_old_cate[0], list_cate[i], list_cate[k]]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)
        wb.save(self.path_to_write)



    def crawlingXiangbao(self, response):
        """箱包"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'箱包'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'箱包')

        for idx,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            # 首页显示类别
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [c['cat_name'].decode('utf-8') for c in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
                list_to_write = [idx+1, list_old_cate[0], '', kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)

            # 扩展(及隐藏)类别
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_cate = [d['cat_name'].decode('utf-8') for d in list_extra_cate]
            list_istitle = [d['is_title']=='true' for d in list_extra_cate]
#            print repr(list_cate).decode('unicode-escape')
#            print list_istitle
            list_index = np.argwhere(np.array(list_istitle)).T[0].tolist()
            list_index.append(len(list_cate))
            list_zip_cate_kw = [(list_index[i], list_index[i+1]) \
                    for i in range(len(list_index)-1)]
            for i,j in list_zip_cate_kw:
                for k in range(i+1,j):
                    list_to_write = [idx+1, list_old_cate[0], list_cate[i], list_cate[k]]
#                    print repr(list_to_write).decode('unicode-escape')
                    ws.append(list_to_write)

        wb.save(self.path_to_write)


    def crawlingXie(self, response):
        """鞋靴"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'鞋靴'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'鞋靴')

        for i,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            list_old_cate = eval(sel.xpath("textarea[1]/text()").extract()[0].strip())
            list_old_cate = [c['cat_name'].decode('utf-8') for c in list_old_cate]
#            print repr(list_old_cate).decode('unicode-escape')
            for kw in list_old_cate[1:]:
#                print list_old_cate[0], kw
                ws.append([2*i+1, list_old_cate[0], kw])

            # 每一大类的扩展类单独作为一个大类
            list_extra_cate = eval(sel.xpath("textarea[2]/text()").extract()[0].strip())
            list_extra_cate = [c['cat_name'].decode('utf-8') for c in list_extra_cate]
#            print repr(list_extra_cate).decode('unicode-escape')
            for kw in list_extra_cate[1:]:
#                print list_extra_cate[0], kw
                ws.append([2*i+2, list_extra_cate[0], kw])
        wb.save(self.path_to_write)


    def crawlingNeiyi(self, response):
        """内衣"""
#        print 'hello alas'

#        with open('data/t.html', 'w') as f:
#            f.write(response.body)

	"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'内衣'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'内衣')

        for i,sel in enumerate(response.xpath("//ul[@class='list-wrap']/li")):
            c1 = sel.xpath("p/a/text()").extract()
            list_kws_c1 = sel.xpath("dl/dd/a/text()").extract()
            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)
	"""


    def crawlingNanzhuang(self, response):
        """男装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'男装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'男装')

        for i,sel in enumerate(response.xpath("//dl[@class='theme-bd-level2']")):
            c1 = sel.xpath("dt/div/a/text()").extract()[0]
            list_kws_c1 = sel.xpath("dd/a/text()").extract()
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)


    def crawlingNvzhuang(self, response):
        """女装"""
        wb = load_workbook(self.path_to_write)
        try:
            wb.remove_sheet(wb[u'女装'])
        except Exception, e:
            pass
        ws = wb.create_sheet(title=u'女装')

        for i,sel in enumerate(response.xpath("//ul[@class='list-wrap']/li")):
            c1 = sel.xpath("p/a/text()").extract()[0]
            list_kws_c1 = sel.xpath("dl/dd/a/text()").extract()
#            print c1, repr(list_kws_c1).decode('unicode-escape')
            for kw in list_kws_c1:
                list_to_write = [i+1, c1, kw]
#                print repr(list_to_write).decode('unicode-escape')
                ws.append(list_to_write)
        wb.save(self.path_to_write)



